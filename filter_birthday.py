#!/usr/bin/env python3

import xlsxwriter
import openpyxl
from openpyxl import Workbook
import re

date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$') 

input_file = 'Data.xlsx'
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = 'filtered Data'

for col_num, header in enumerate(['STT', 'Name', 'Birthday', 'Hometown'], start=1):
    new_sheet.cell(row=1, column=col_num, value=header)

new_row = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    STT, Name, Birthday, Hometown = row
    if isinstance(Birthday, str) and date_pattern.match(Birthday): 
        new_sheet.cell(row = new_row, column=1, value=STT)
        new_sheet.cell(row = new_row, column=2, value=Name)
        new_sheet.cell(row = new_row, column=3, value=Birthday)
        new_sheet.cell(row = new_row, column=4, value=Hometown)
        new_row += 1

output_file='filtered_data.xlsx'
new_wb.save(output_file)
