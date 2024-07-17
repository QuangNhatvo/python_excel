#!/usr/bin/env pyhthon3

import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
from random import randint, choice
from datetime import datetime, timedelta

workbook = openpyxl.Workbook()

sheet = workbook.active
sheet.title = "Data"

headers = ['STT', 'Name', 'Birthday', 'Hometown']
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column = col_num, value=header)

num_rows = 50
start_date = datetime(1980,1,1)
end_date = datetime (2005,12,31)

for i in range(2, num_rows + 2):
    sheet.cell(row=i, column=1, value=i-1)
    sheet.cell(row=i, column=2, value= choice(['Nguyen van Aaaa', 'Tran Thi Bbb', 'Hoang gia ccc', 'Tran Van Dddd','nguyen thu ttttt', 'hoang van Hhhh', 'Doan van Uuuu' ]))
    random_date = start_date + timedelta(days=randint(0,(end_date - start_date).days))
    sheet.cell(row=i, column=3, value=random_date.strftime('%Y-%m-%d').replace(
            '-0','-' ))
    sheet.cell(row=i, column=4, value= choice(['Ha Noi', 'Ho Chi Minh', 'Hue', 'Da Nang', 'Hai Phong', 'Can Tho', 'Dong Nai', 'Binh Duong', 'Vung Tau', 'Long An', 'Tra Vinh']))

file_name = "Data.xlsx"
workbook.save(file_name)
